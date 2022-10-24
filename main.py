from openpyxl import load_workbook, Workbook
from read_url import parse_url
from read_data import copy_template, search_start_point, search_data_area, add_data, write_to_output

input_file = 'TestTask_1_input_1.xlsx'
template_file = 'TestTask_1_output_blank_template.xlsx'
output_file = 'TestTask_1_output_1.xlsx'

try:
    # открываем таблицу с исходными данными
    wb_inp = load_workbook(input_file)
    sheet = wb_inp.active
except FileNotFoundError as FNF:
    print(FNF)
    print('Check path to input file or add file to directory with script')

# в этот список добавляем все исходные данные и данные полученные по ссылкам
extracted_data = []

# создаем новую книгу(объект книги), удаляем из книги активную стандартную таблицу,
# создаем свою таблицу out_sheet, сохраняем в файл (.xlsx) и закрываем
output_book = Workbook()
output_book.remove(output_book.active)
out_sheet = output_book.create_sheet('out_sheet')


try:
    # загружаем таблицу с шаблоном
    wb_out_temp = load_workbook(template_file)
    sht_temp = wb_out_temp.active
except FileNotFoundError as FNF:
    print(FNF)
    print('Check path to template file or add file to directory with script')

# копируем шаблон и размещаем в известный диапазон
copy_template(sht_temp, out_sheet)

# находим ячейки, где располагается начало таблицы с исходными данными
table_starts = search_start_point(sheet)
# выполняем поиск исходных данных и добавляем их в список extracted_data
for i in table_starts:
    data_area = search_data_area(i, sheet)
    add_data(extracted_data, data_area, sheet)

for i in extracted_data:
    # три попытки выполнить парсинг сайта на случай ошибок или нестабильного соединения
    tries = 3
    while tries > 0:
        try:
            vendor_code, price = parse_url(i[1])
            i.append(vendor_code)
            i.append(price)
            tries = 0
        except ValueError as VEex:
            print('Data not received')
            tries -= 1
            continue
        except Exception as ex:
            print(ex)
            tries -= 1
            continue
        finally:
            if len(i) < 3 and tries == 0:
                i.append('Not found')
                i.append('---')
    print(i)

# записываем данные в выходной эксель-файл
write_to_output(extracted_data, out_sheet)
# флаг позволяющий повторить попытку записи результатов в сводную таблицу
writing_result = True
while writing_result:
    try:
        output_book.save(output_file)
        writing_result = False
    except PermissionError as PE:
        print(PE)
        action_confirmation = input('Close the PivotTable file if it has already been created.\n'
                                    'Enter "Y" if the file is closed and you want to retry writing.\n'
                                    'Enter anything to close the program.\n')
        if action_confirmation.lower() == 'y':
            continue
        else:
            writing_result = False
    finally:
        output_book.close()
        wb_inp.close()
        wb_out_temp.close()
