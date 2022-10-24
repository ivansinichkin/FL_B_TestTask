# coding=utf-8
from openpyxl.styles import Alignment
from copy import copy


def copy_template(sheet_from, sheet_to):
    sheet_to.merged_cells = copy(sheet_from.merged_cells)
    row_shift = sheet_from.min_row - 2
    col_shift = sheet_from.min_column - 2
    # -2, чтобы таблица начиналась с ячейки B2
    for row in sheet_from.iter_rows(min_row=sheet_from.min_row, min_col=sheet_from.min_column,
                                    max_row=sheet_from.max_row, max_col=sheet_from.max_column):
        for cell in row:
            row_to = cell.row - row_shift
            col_to = cell.column - col_shift
            # копируем содержимое, стиль, и горизонтальное выравнивание ячеек
            sheet_to.cell(row=row_to, column=col_to).value = cell.value
            if cell.has_style:
                sheet_to.cell(row=row_to, column=col_to).border = copy(cell.border)  # работает
            # sheet_to.cell(row=row_to, column=col_to).border = copy(cell.border)  # тоже работает
            hor_align = cell.alignment.horizontal
            sheet_to.cell(row=row_to, column=col_to).alignment = Alignment(horizontal=hor_align)


def search_start_point(sht):
    """
    Функция выполняет поиск ячеек в которых содержится запись "Данные по товарам", эту ячейку считаем
    левым верхним углом таблицы с данными.
    :param sht: sheet object in workbook
    :return coordinates_list: list of tuples with coordinates cell which contain 'Данные по товарам'
    ex.(row_index, col_index)
    """
    coordinates_list = []
    for row in sht.iter_rows(min_row=sht.min_row, min_col=sht.min_column,
                             max_row=sht.max_row, max_col=sht.max_column):
        for cell in row:
            if cell.value == 'Данные по товарам':
                coordinates = (cell.row, cell.column - 1)  # вычитаем единицу, тк при обращении к ячейке
                # с индексом без вычитания попадаем в ячейку правее
                coordinates_list.append(coordinates)
    return coordinates_list


def search_data_area(start_cell, sheet):
    """
    Функция получает на вход координату ячейки "Данные по товарам", а возвращает координаты диапазона ячеек,
    где находятся непосредтсвенно данные
    Эта версия функции учитвает, что в таблице может быть строка с незаполненными данными, а ниже с заполненными
    :param start_cell: список из двух параметров, который указывает на ячейку с записью "Данные по товарам"
    :param sheet: таблица с исходными данными
    :return : кортеж координат вида (номер ряда ВЛ угла, номер колонны ВЛ угла,
    номер ряда НП угла, номер колонны НП угла
    """
    none_counter_in_row = 0
    # следующее смещение индексов нужно для того, чтобы начать поиск данных из ячейки с первым наименованием
    row_index = start_cell[0] + 2
    col_index = start_cell[1] + 1
    stop_cell = ()
    # проход по рядам
    while isinstance(sheet[row_index][col_index - 1].value, int):
        if sheet[row_index][col_index].value is None:
            row_index += 1
            col_index = start_cell[1] + 1
            none_counter_in_row = 0
            continue
        # проход по ряду пока не найдется второй None
        while none_counter_in_row < 2:
            if sheet[row_index][col_index].value is None:
                none_counter_in_row += 1
            stop_cell = (row_index, col_index - 1)  # эта единица нужна для того, чтобы оставить
            # границу в ячейке с содержимым, а не в соседней, где уже None
            col_index += 1
        row_index += 1
        col_index = start_cell[1] + 1
        none_counter_in_row = 0
    return start_cell[0] + 2, start_cell[1] + 1, stop_cell[0], stop_cell[1]


def add_data(data_list, table_area, sheet):
    """
    функция записывает данные в список, получая на вход диапазон ячеек в котором содержатся данные
    :param data_list: список, в который записываются все необходимые данные
    :param table_area: диапазон ячеек, в котором содержится таблица с данными
    :param sheet: таблица исходных данных
    """
    for row in sheet.iter_rows(min_row=table_area[0], min_col=table_area[1] + 1,
                               max_row=table_area[2], max_col=table_area[3] + 1):  # добавляем единицы к индексам
        # столбцов потому что здесь столбцы начинаются с 1
        extracted_data_row = []
        for cell in row:
            if cell.value is None:
                continue
            extracted_data_row.append(cell.value)
        if len(extracted_data_row) > 1:
            data_list.append(extracted_data_row)


def write_to_output(recieved_data, out_sht):
    counter = 1
    cur_row = 4
    for i in recieved_data:
        # добавим границы и выравнивание к каждой записи, как к первой
        # этот цикл нужен только для установки границ и выравнивания
        for m in range(2, 8):
            out_sht.cell(row=cur_row, column=m).border = copy(out_sht.cell(row=4, column=m).border)
            hor_align = out_sht.cell(row=4, column=m).alignment.horizontal
            out_sht.cell(row=cur_row, column=m).alignment = Alignment(horizontal=hor_align)

        out_sht.cell(row=cur_row, column=2).value = counter  # номер
        out_sht.cell(row=cur_row, column=3).value = i[0]  # наименование
        out_sht.cell(row=cur_row, column=6).value = i[1]  # ссылка
        if len(i) == 4:
            out_sht.cell(row=cur_row, column=5).value = i[2]  # артикул
            out_sht.cell(row=cur_row, column=7).value = i[3]  # цена
        counter += 1
        cur_row += 1
